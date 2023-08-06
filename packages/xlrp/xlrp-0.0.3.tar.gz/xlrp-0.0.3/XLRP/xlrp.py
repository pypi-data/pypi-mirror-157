import time
import traceback
from functools import wraps, reduce
from XLRP.drawplot import DrawPlot
from pathlib import Path
import os
from XLRP.xlconf import XlConf
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


_test_result = {"sys": '', "model": [], 'step': []}
_error = {}

class SysName:
    """
    获取系统名称的类
    """
    def __init__(self, sys_name):
        self.sys_name = sys_name

    def __enter__(self):
        pass

    def __exit__(self, exc_type, exc_val, exc_tb):
        _test_result["sys"] = self.sys_name

    def __call__(self, cls):
        _test_result['sys'] = self.sys_name

        @wraps(cls)
        def decorator(*args, **kwargs):
            res = cls(*args, **kwargs)
            return res
        return decorator

class ModelName:
    """
    获取模块名称类
    """
    def __init__(self, model_name):
        self.model = model_name

    def __enter__(self):
        ...

    def __exit__(self, exc_type, exc_val, exc_tb):
        _test_result['model'].append(self.model)

    def __call__(self, func):
        @wraps(func)
        def decorator(*args, **kwargs):
            _test_result["model"].append(self.model)
            res = func(*args, **kwargs)
            return res
        return decorator

class StepName:
    """
    获取step步骤(用例)的类
    """
    def __init__(self, step_name):
        self.step = step_name

    def __enter__(self):
        self.start_time = time.time()

    def __exit__(self, exc_type, exc_val, exc_tb):
        if any([exc_type, exc_val, exc_tb]):
            # error_str = str([''.join(x) for x in traceback.format_exception(exc_type, exc_val, exc_tb)][0])
            error_str = reduce(lambda x, y: x + y, traceback.format_exception(exc_type, exc_val, exc_tb))
            _error[self.step] = error_str
            if exc_type is AssertionError:
                case_status = False
            else:
                case_status = 'Error'
        else:
            case_status = True
            error_str = None
        stop_time = time.time()
        run_time = round(stop_time - self.start_time, 3)
        _test_result["step"].append({self.step: {'run_time': run_time, 'status': case_status, 'msg': error_str}})

    def __call__(self, func):
        step = self.step

        @wraps(func)
        def decorator(*args, **kwargs):
            start_time = time.time()
            try:
                res = func(*args, **kwargs)
                if type(res) is bool:
                    case_status = res
                else:
                    case_status = True
                error_str = None
            except Exception as e:
                case_status = False
                res = None
                # error_str = str([''.join(x) for x in traceback.format_exc()][0])
                error_str = reduce(lambda x, y: x + y, traceback.format_exc())
            stop_time = time.time()
            run_time = round(stop_time - start_time, 3)
            _test_result["step"].append({step: {"run_time": run_time, "status": case_status, 'msg': error_str}})
            return res
        return decorator

class _ProcessingData:
    def __init__(self, data_list: dict):
        self.data_list = data_list
        self.model_list = self.data_list['model']
        self.filter_models = []
        [self.filter_models.append(x) for x in self.model_list if x not in self.filter_models]
        self.steps = self.data_list['step']
        self.zips_data = list(zip(self.model_list, self.steps))

    def pie_data(self):
        """
        修整饼图数据
        :return:
        """
        passed = 0
        failed = 0
        error = 0
        for dc in self.steps:
            status = list(dc.values())[0]['status']
            if status is True:
                passed += 1
            elif status == 'Error':
                error += 1
            elif status is False:
                failed += 1
            else:
                failed += 1
        return passed, failed, error

    def bar_data(self):
        """
        修整柱状图的测试数据
        :return:
        """
        dc_data = {}
        for model in self.filter_models:
            dc_data[model] = {'passed': 0, 'failed': 0, 'error': 0}
            for case in self.zips_data:
                if model == case[0]:
                    status = list(case[1].values())[0]['status']
                    if status is True:
                        dc_data[model]['passed'] += 1
                    elif status is False:
                        dc_data[model]['failed'] += 1
                    elif status == 'Error':
                        dc_data[model]['error'] += 1
                    else:
                        dc_data[model]['failed'] += 1
        return dc_data

    def plot_data(self):
        """
        修正绘制折线图的测试数据
        :return:
        """
        dc_data = {}
        for model in self.filter_models:
            dc_data[model] = {}
            for case in self.zips_data:
                if model == case[0]:
                    run_time = {list(case[1].keys())[0]: list(case[1].values())[0]["run_time"]}
                    dc_data[model].update(run_time)
        return dc_data

class Runner:
    def __init__(self, show_plot=False):
        self.show_plot = show_plot
        self.save_path = Path().cwd() / '.xlrp_plot'
        self.save_path.mkdir(exist_ok=True)
        self.save_path = os.fspath(self.save_path)

    def run(self, obj, param_iter=None):
        """
        运行单个用例的方法
        :param obj: 用例函数或者方法名
        :param param_iter: 参数，使用同parameter，需要列表或者元组形式，会循环传入到函数或者方法中
        :return: self
        """
        if type(param_iter) not in (list, tuple):
            if param_iter is None:
                try:
                    obj()
                except:
                    ...
            else:
                raise ValueError("The parameters must be lists or tuples")
        else:
            for param in param_iter:
                try:
                    if type(param) in (list, tuple):
                        obj(*param)
                    elif type(param) is dict:
                        obj(**param)
                    else:
                        obj()
                except:
                    ...
        return self

    def run_class(self, test_class):
        """
        运行整个Class方法，仅收集包含'xl'开头的方法
        :param test_class: 需要进行运行的类名
        :return: self
        """
        class_dirs = test_class().__dir__()
        print(class_dirs)
        func_list = [x for x in class_dirs if x.startswith("xl")]
        for func in func_list:
            print(func)
            try:
                eval(f'test_class().{func}()')
            except:
                ...
        return self

    def plot_save_excel(self, file_path=None, width_list=(500, 500, 500)):
        """
        将plot图片保存到excel的代码
        :param file_path: excel文件路径
        :param width_list: 图片的宽度(默认第三个数据，会在原基础上X2，改动在xlconf文件中)
        :return: self
        """
        sys_name = _test_result.get('sys')
        prd = _ProcessingData(_test_result)
        pie_data = prd.pie_data()
        bar_data = prd.bar_data()
        plot_data = prd.plot_data()
        dp = DrawPlot(sys_name=sys_name, save_path=self.save_path)
        pie_path = dp.draw_pie(通过=pie_data[0], 失败=pie_data[1], 错误=pie_data[2], show_plot=self.show_plot)
        bar_path = dp.draw_bar(**bar_data, show_plot=self.show_plot)
        plot_path = dp.draw_plot(**plot_data, show_plot=self.show_plot)
        pic_list = [pie_path, bar_path, plot_path]
        if file_path is not None:
            # dp.plot_to_excel(file_path=file_path, pic_list=pic_list, width_list=width_list, cells=cells)
            xc = XlConf(filepath=file_path, sheetname='TestReport', pic_list=pic_list, pic_width=width_list)
            xc.set_cell_size(create_new=True)
            del xc
            xc2 = XlConf(filepath=file_path, sheetname='TestReport', pic_list=pic_list, pic_width=width_list)
            xc2.save_plot()

    def save_default(self):
        sys_name = _test_result.get('sys')
        excel_name = f'{sys_name}系统测试EXCEL报告.xlsx'
        report_path = Path().cwd() / 'XlrpReport'
        report_path.mkdir(exist_ok=True)
        model_list = _test_result['model']
        steps = _test_result['step']
        # 创建excel表格存入数据
        wb = openpyxl.Workbook()
        ws = wb.create_sheet('TestCases')
        title_font = Font(size=16, name='微软雅黑', bold=True)
        label_font = Font(size=14, name='微软雅黑')
        content_font = Font(size=11, name='微软雅黑', bold=True)
        false_font = Font(size=11, name='微软雅黑', color='cc3333', bold=True)
        title_fill = PatternFill(fgColor='adc2eb', fill_type='solid')
        label_fill = PatternFill(fgColor='b3b3b3', fill_type='solid')
        # false_fill = PatternFill(fgColor='cc3333', fill_type='solid')
        cell_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.merge_cells('A1:F1')
        ws['A1'].value = f'{sys_name}系统运行用例情况表'
        ws['A1'].font = title_font
        ws['A1'].fill = title_fill
        ws['A1'].alignment = cell_align
        labels = ['编号', '模块', '用例名称', '运行时间', '运行结果', '异常信息']
        for index, label in enumerate(labels):
            label_cell = ws.cell(row=2, column=index + 1)
            label_cell.value = label
            label_cell.font = label_font
            label_cell.fill = label_fill
            label_cell.alignment = cell_align
        for index, case_data in enumerate(zip(model_list, steps)):
            model = case_data[0]
            step = case_data[1]
            case_name = list(step.keys())[0]
            runtime = step[case_name]['run_time']
            status_origin = step[case_name]['status']
            if status_origin is True:
                status = 'PASSED'
            elif status_origin is False:
                status = 'FAILED'
            else:
                status = 'ERROR'
            msg = step[case_name]['msg']
            data = [str(index + 1), model, case_name, runtime, status, msg]
            for col, param in enumerate(data):
                case_cell = ws.cell(row=index + 3, column=col + 1)
                case_cell.value = param
                if status != 'PASSED':
                    case_cell.font = false_font
                else:
                    case_cell.font = content_font
                case_cell.alignment = cell_align
                # if status is not True:
                #     case_cell.fill = false_fill
        sava_path = os.fspath(report_path) + '/' + excel_name
        wb.save(sava_path)
        wb.close()
        del wb
        wb2 = openpyxl.load_workbook(sava_path)
        ws2 = wb2['TestCases']
        try:
            wb2.remove(wb2['Sheet'])
        except:
            ...
        for col in range(1, len(labels) + 1):
            width = 25
            if col == len(labels):
                width = 50
            ws2.column_dimensions[get_column_letter(col)].width = width
        old_model = ''
        for row in range(3, ws2.max_row + 1):
            current_model = ws2.cell(row=row, column=2).value
            if old_model == current_model:
                ws2.merge_cells(f'B{row - 1}:B{row}')
            else:
                old_model = current_model
        wb2.save(sava_path)
        wb2.close()
        del wb2
        self.plot_save_excel(file_path=sava_path)

@SysName("测试模块")
class ExcelTest:
    @ModelName('模块1')
    def xl_pr(self, a, b, step):
        with StepName(step):
            assert a == b

    @ModelName('模块2')
    def xl_pr2(self, a, b, step):
        with StepName(step):
            assert a > b/a

    @ModelName("模块3")
    @StepName("步骤")
    def xl_pr3(self):
        print(123)

    @ModelName("模块4")
    @StepName("步骤")
    def xl_pr4(self):
        print(666)

if __name__ == '__main__':
    # run(ExcelTest().pr, [[1, 2, "步骤1"], [3, 3, "步骤2"], [5, 1, "步骤3"]], show_plot=True)
    # run(ExcelTest().pr2, [[1, 2, "步骤1"], [3, 3, "步骤2"], [0, 4, "步骤3"]], show_plot=True)
    # run(ExcelTest().pr3, show_plot=True)
    runner = Runner()
    runner.run(ExcelTest().xl_pr, [[1, 2, "步骤1"], [3, 3, "步骤2"], [5, 1, "步骤3"]])
    runner.run(ExcelTest().xl_pr2, [[1, 2, "步骤1"], [3, 3, "步骤2"], [0, 4, "步骤3"]])
    runner.run(ExcelTest().xl_pr3)
    runner.save_default()
    print(_test_result)
    # runner.plot_save_excel('./test.xlsx')
    # runner.run_class(ExcelTest)
    # runner.plot_save_excel('./test.xlsx')